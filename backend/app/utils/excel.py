"""Excel (.xlsx) hisobotlarini tayyorlash.

Do'kondor ma'lumotlarini o'zida saqlab qo'yishi, buxgalteriyaga berishi
yoki tizimdan chiqib ketsa ham arxivi qolishi uchun.
"""
import io
import logging
from datetime import datetime
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import utcnow
from app.utils.helpers import to_local, debt_status_label

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1E3A8A")
_MONEY_FMT = '#,##0" so\'m"'
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Muddati o'tgan qatorlarni ajratib ko'rsatish
_OVERDUE_FILL = PatternFill("solid", fgColor="FEE2E2")
_CLOSED_FILL = PatternFill("solid", fgColor="ECFDF5")


def _dt(value: Optional[datetime]) -> str:
    """Sanani Toshkent vaqtida, Excel o'qiy oladigan ko'rinishda."""
    if not value:
        return "—"
    return to_local(value).strftime("%d.%m.%Y %H:%M")


def _date(value: Optional[datetime]) -> str:
    if not value:
        return "Muddatsiz"
    return to_local(value).strftime("%d.%m.%Y")


def _sheet_name(title: str) -> str:
    """Excel varaq nomi: 31 belgi, `: \\ / ? * [ ]` belgilari taqiqlangan."""
    for ch in ':\\/?*[]':
        title = title.replace(ch, "-")
    return title.strip()[:31] or "Hisobot"


def _write_sheet(ws, title: str, headers: list[str], rows: Iterable[list], widths: list[int]):
    ws.title = _sheet_name(title)

    ws.append([title])
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append([f"Yuklangan: {_dt(utcnow())} (Toshkent)"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row in rows:
        ws.append(row)

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Sarlavha qatorini muzlatamiz — uzun ro'yxatda qulay
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"
    return header_row


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Do'kon hisoboti (qarzdorlar + qarzlar) ───────────────────────────────────

def build_shop_report(shop_name: str, clients: list[dict], debts: list[dict]) -> bytes:
    """Ikki varaqli hisobot: «Qarzdorlar» va «Qarzlar»."""
    wb = Workbook()

    # 1-varaq: qarzdorlar xulosasi
    ws = wb.active
    header_row = _write_sheet(
        ws,
        f"{shop_name} — qarzdorlar",
        ["№", "Mijoz", "Telefon", "Faol qarzlar", "Qoldiq", "Jami to'langan", "Holat"],
        [
            [
                i,
                c["full_name"],
                c["phone"],
                c["active_debts"],
                c["total_remaining"],
                c["total_paid"],
                "Muddati o'tgan" if c["has_overdue"] else ("Qarzsiz" if c["total_remaining"] == 0 else "Faol"),
            ]
            for i, c in enumerate(clients, start=1)
        ],
        [5, 28, 18, 14, 18, 18, 16],
    )
    for row in range(header_row + 1, ws.max_row + 1):
        ws.cell(row=row, column=5).number_format = _MONEY_FMT
        ws.cell(row=row, column=6).number_format = _MONEY_FMT
        if ws.cell(row=row, column=7).value == "Muddati o'tgan":
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = _OVERDUE_FILL

    total_remaining = sum(c["total_remaining"] for c in clients)
    ws.append([])
    ws.append(["", "JAMI", "", "", total_remaining, sum(c["total_paid"] for c in clients), ""])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).number_format = _MONEY_FMT
    ws.cell(row=ws.max_row, column=6).number_format = _MONEY_FMT

    # 2-varaq: qarzlar tafsiloti
    ws2 = wb.create_sheet()
    header_row2 = _write_sheet(
        ws2,
        "Qarzlar",
        ["Raqam", "Mijoz", "Telefon", "Summa", "To'langan", "Qoldiq",
         "Holat", "Muddat", "Izoh", "Yaratilgan"],
        [
            [
                d["debt_number"], d["client_name"], d["client_phone"],
                d["amount"], d["paid_amount"], d["remaining"],
                debt_status_label(d["status"]), _date(d["due_date"]),
                d["note"] or "", _dt(d["created_at"]),
            ]
            for d in debts
        ],
        [12, 26, 17, 16, 16, 16, 18, 14, 30, 18],
    )
    for row in range(header_row2 + 1, ws2.max_row + 1):
        for col in (4, 5, 6):
            ws2.cell(row=row, column=col).number_format = _MONEY_FMT
        status = ws2.cell(row=row, column=7).value
        if status == debt_status_label("overdue"):
            fill = _OVERDUE_FILL
        elif status == debt_status_label("closed"):
            fill = _CLOSED_FILL
        else:
            continue
        for col in range(1, 11):
            ws2.cell(row=row, column=col).fill = fill

    return _save(wb)


# ─── Admin: do'konlar ro'yxati ────────────────────────────────────────────────

def build_shops_report(shops: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    header_row = _write_sheet(
        ws,
        "Do'konlar ro'yxati",
        ["№", "Do'kon", "Egasi", "Telefon", "Holat", "Mijozlar",
         "Faol qarzlar", "Trial tugashi", "Obuna tugashi", "Ro'yxatdan o'tgan"],
        [
            [
                i, s["name"], s["owner"], s["owner_phone"], s["status"],
                s["client_count"], s["active_debts"],
                _date(s["trial_end"]), _date(s["subscription_end"]), _dt(s["created_at"]),
            ]
            for i, s in enumerate(shops, start=1)
        ],
        [5, 26, 24, 18, 12, 11, 14, 15, 15, 18],
    )
    del header_row
    return _save(wb)
